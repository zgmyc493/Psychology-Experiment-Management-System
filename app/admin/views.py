from flask import render_template, redirect, url_for, flash, request, send_file, current_app, jsonify
from flask_login import login_required, current_user
from app import db
from app.models.user import User
from app.models.experiment import Experiment, Participation
from app.models.notification import Notification
from . import admin
from functools import wraps
from datetime import datetime, timedelta
import openpyxl
from io import BytesIO
from flask_wtf import FlaskForm
from sqlalchemy import or_, and_, cast, Integer, extract

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            # 如果用户未登录，重定向到登录页面，并记录下一个页面的URL
            return redirect(url_for('auth.login', next=request.url))
        if current_user.role != 'admin':
            flash('您没有权限访问此页面。', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated_function

@admin.route('/')
@login_required
@admin_required
def index():
    """管理员首页"""
    # 获取用户统计
    total_all_users = User.query.count()
    participant_count = User.query.filter_by(role='participant').count()
    admin_count = User.query.filter_by(role='admin').count()
    active_users = User.query.filter_by(is_active=True).count()
    inactive_users = User.query.filter_by(is_active=False).count()
    
    # 获取今日和昨日新增用户数
    today = datetime.utcnow().date()
    today_users = User.query.filter(
        db.func.date(User.created_at) == today
    ).count()
    yesterday = today - timedelta(days=1)
    yesterday_users = User.query.filter(
        db.func.date(User.created_at) == yesterday
    ).count()
    user_increase = f"+{today_users - yesterday_users}" if today_users > yesterday_users else str(today_users - yesterday_users)
    
    # 获取实验统计
    total_experiments = Experiment.query.count()
    draft_experiments = Experiment.query.filter_by(status='draft').count()
    active_experiments = Experiment.query.filter_by(status='active').count()
    completed_experiments = Experiment.query.filter_by(status='completed').count()
    cancelled_experiments = Experiment.query.filter_by(status='cancelled').count()
    
    # 获取参与情况统计
    total_participations = Participation.query.count()
    pending_participations = Participation.query.filter_by(status='pending').count()
    completed_participations = Participation.query.filter_by(status='completed').count()
    completion_rate = round((completed_participations / total_participations * 100) if total_participations > 0 else 0)
    
    return render_template('admin/index.html',
                         active_page='dashboard',
                         total_all_users=total_all_users,
                         participant_count=participant_count,
                         admin_count=admin_count,
                         active_users=active_users,
                         inactive_users=inactive_users,
                         today_users=today_users,
                         user_increase=user_increase,
                         total_experiments=total_experiments,
                         draft_experiments=draft_experiments,
                         active_experiments=active_experiments,
                         completed_experiments=completed_experiments,
                         cancelled_experiments=cancelled_experiments,
                         total_participations=total_participations,
                         pending_participations=pending_participations,
                         completed_participations=completed_participations,
                         completion_rate=completion_rate)

@admin.route('/users')
@login_required
@admin_required
def user_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    gender = request.args.get('gender', '')
    age_range = request.args.get('age_range', '')
    graduation_year = request.args.get('graduation_year', '')
    status = request.args.get('status', '')
    school = request.args.get('school', '')
    college = request.args.get('college', '')

    # 构建基础查询，普通管理员不能查看超级管理员
    query = User.query.filter(User.role != 'superadmin')

    # 搜索条件
    if search:
        query = query.filter(
            or_(
                User.username.ilike(f'%{search}%'),
                User.email.ilike(f'%{search}%')
            )
        )

    # 性别筛选
    if gender:
        query = query.filter(User.gender == gender)

    # 年龄范围筛选
    if age_range:
        age_ranges = {
            '18-22': (18, 22),
            '23-26': (23, 26),
            '27-30': (27, 30),
            '30+': (30, 200)  # 使用一个较大的数作为上限
        }
        if age_range in age_ranges:
            min_age, max_age = age_ranges[age_range]
            query = query.filter(User.age >= min_age, User.age <= max_age)

    # 毕业年份筛选
    if graduation_year:
        query = query.filter(extract('year', User.graduation_date) == int(graduation_year))

    # 学校筛选
    if school:
        query = query.filter(User.school.ilike(f'%{school}%'))

    # 学院筛选
    if college:
        query = query.filter(User.college.ilike(f'%{college}%'))

    # 状态筛选
    if status:
        is_active = status == 'active'
        query = query.filter(User.is_active == is_active)

    # 分页
    users = query.paginate(page=page, per_page=10)
    form = FlaskForm()  # 创建一个空的表单对象，用于CSRF令牌

    # 获取当前年份，用于毕业年份选项
    current_year = datetime.now().year

    return render_template('admin/user_list.html', 
                         active_page='users',
                         users=users, 
                         form=form,
                         current_year=current_year)

@admin.route('/user/<int:id>/toggle_status', methods=['POST'])
@login_required
@admin_required
def toggle_user_status(id):
    user = User.query.get_or_404(id)
    # 检查是否是超级管理员
    if user.role == 'superadmin':
        flash('无权修改超级管理员的状态。', 'danger')
        return redirect(url_for('admin.user_list'))
    
    if user == current_user:
        flash('不能修改自己的状态。', 'warning')
    else:
        user.is_active = not user.is_active
        db.session.commit()
        status = '启用' if user.is_active else '禁用'
        flash(f'用户 {user.username} 已{status}。', 'success')
    return redirect(url_for('admin.user_list'))

@admin.route('/user/<int:id>/change_role', methods=['POST'])
@login_required
@admin_required
def change_user_role(id):
    user = User.query.get_or_404(id)
    # 检查是否是超级管理员
    if user.role == 'superadmin':
        flash('无权修改超级管理员的角色。', 'danger')
        return redirect(url_for('admin.user_list'))
    
    if user == current_user:
        flash('不能修改自己的角色。', 'warning')
    else:
        new_role = request.form.get('role')
        if new_role in ['admin', 'participant']:  # 确保不能将用户设置为超级管理员
            old_role = user.role
            user.role = new_role
            
            # 创建角色变更通知
            role_display = '管理员' if new_role == 'admin' else '普通用户'
            notification = Notification(
                user_id=user.id,
                title='用户角色变更',
                message=f'您的用户角色已被更改为{role_display}。',
                type='role_changed'
            )
            db.session.add(notification)
            db.session.commit()
            
            flash(f'用户 {user.username} 的角色已更改为 {new_role}。', 'success')
    return redirect(url_for('admin.user_list'))

@admin.route('/experiments')
@login_required
@admin_required
def experiment_list():
    """实验列表页面"""
    page = request.args.get('page', 1, type=int)
    experiments = Experiment.query.filter(
        Experiment.deleted_at.is_(None)  # 只显示未删除的实验
    ).order_by(
        Experiment.created_at.desc()
    ).paginate(
        page=page, 
        per_page=10,
        error_out=False
    )
    return render_template('admin/experiment_list.html', 
                         active_page='experiments',
                         experiments=experiments,
                         timedelta=timedelta)

@admin.route('/settings')
@login_required
@admin_required
def settings():
    """系统设置页面"""
    return render_template('admin/settings.html',
                         active_page='settings')

@admin.route('/export')
@login_required
@admin_required
def export_data():
    """数据导出页面"""
    return render_template('admin/export.html')

@admin.route('/export_users')
@login_required
@admin_required
def export_users():
    # 创建新的Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用户数据"
    
    # 设置表头
    headers = [
        'ID', '用户名', '邮箱', '角色', '状态', '注册时间', '最后登录',
        '性别', '年龄', '学校', '学院', '毕业时间', '电话号码'
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 写入用户数据，排除超级管理员
    users = User.query.filter(User.role != 'superadmin').all()
    for row, user in enumerate(users, 2):
        # 基本信息
        ws.cell(row=row, column=1, value=user.id)
        ws.cell(row=row, column=2, value=user.username)
        ws.cell(row=row, column=3, value=user.email)
        ws.cell(row=row, column=4, value='管理员' if user.role == 'admin' else '普通用户')
        ws.cell(row=row, column=5, value='启用' if user.is_active else '禁用')
        ws.cell(row=row, column=6, value=user.local_created_at.strftime('%Y-%m-%d %H:%M') if user.local_created_at else '')
        ws.cell(row=row, column=7, value=user.local_last_login.strftime('%Y-%m-%d %H:%M') if user.local_last_login else '从未登录')
        
        # 个人信息
        ws.cell(row=row, column=8, value='男' if user.gender == 'male' else '女' if user.gender == 'female' else '')
        ws.cell(row=row, column=9, value=user.age)
        ws.cell(row=row, column=10, value=user.school)
        ws.cell(row=row, column=11, value=user.college)
        ws.cell(row=row, column=12, value=user.graduation_date.strftime('%Y-%m-%d') if user.graduation_date else '')
        ws.cell(row=row, column=13, value=user.phone)
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # 保存到内存中
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # 生成下载文件名
    filename = f'用户数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@admin.route('/user/<int:id>/profile')
@login_required
@admin_required
def view_user_profile(id):
    """查看用户个人信息"""
    user = User.query.get_or_404(id)
    if user.role == 'admin':
        flash('管理员没有个人信息页面。', 'warning')
        return redirect(url_for('admin.user_list'))
    
    return render_template('admin/user_profile.html', user=user)

@admin.route('/experiment/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_experiment(id):
    """软删除实验"""
    experiment = Experiment.query.get_or_404(id)
    experiment.soft_delete(current_user.id)
    db.session.commit()
    flash(f'实验"{experiment.title}"已移至回收站', 'info')
    return redirect(url_for('admin.experiment_list'))

@admin.route('/experiment/recycle')
@login_required
@admin_required
def experiment_recycle():
    """实验回收站"""
    deleted_experiments = Experiment.query.filter(Experiment.deleted_at.isnot(None)).order_by(Experiment.deleted_at.desc()).all()
    return render_template('admin/experiment_recycle.html', deleted_experiments=deleted_experiments)

@admin.route('/experiment/<int:id>/restore', methods=['POST'])
@login_required
@admin_required
def restore_experiment(id):
    """恢复已删除的实验"""
    experiment = Experiment.query.get_or_404(id)
    if not experiment.is_deleted:
        flash('该实验未被删除', 'warning')
        return redirect(url_for('admin.experiment_recycle'))
    
    # 获取所有参与者
    participants = [p.user for p in experiment.participations.all()]
    
    experiment.restore()
    
    # 为每个参与者创建通知
    for participant in participants:
        notification = Notification(
            user_id=participant.id,
            title='实验已恢复',
            message=f'您参与的实验"{experiment.title}"已被管理员恢复，您可以继续参与该实验。',
            type='experiment_restored'
        )
        db.session.add(notification)
    
    db.session.commit()
    flash(f'实验"{experiment.title}"已恢复', 'success')
    return redirect(url_for('admin.experiment_recycle'))

@admin.route('/experiment/<int:id>/permanent_delete', methods=['POST'])
@login_required
@admin_required
def permanent_delete_experiment(id):
    """永久删除实验"""
    experiment = Experiment.query.get_or_404(id)
    if not experiment.is_deleted:
        flash('请先将实验移至回收站', 'warning')
        return redirect(url_for('admin.experiment_list'))
    
    db.session.delete(experiment)
    db.session.commit()
    flash(f'实验"{experiment.title}"已永久删除', 'success')
    return redirect(url_for('admin.experiment_recycle')) 